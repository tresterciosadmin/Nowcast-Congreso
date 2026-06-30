"""Persistencia de la base de PROYECTOS DE LEY (SQLite) + export a Excel.

Consume la salida del extractor de giros (`datos/seguimiento` → `FichaExpediente`
serializada a dict) y la guarda en una base SQLite, que es la fuente de verdad.
No importa código de otro módulo: recibe un dict (el contrato), no la clase.

Idempotente por DENOMINADOR:
  • re-cargar el mismo proyecto NO duplica: actualiza estado, giros, autores y
    trámite (reflejan el estado oficial actual) y conserva `creado_en`.
  • las TAXONOMÍAS no se tocan acá (las maneja el agente); se preservan entre scrapes.

Uso:
    from store import conectar, upsert_proyecto, export_excel
    con = conectar("data/proyectos.db")
    upsert_proyecto(con, ficha_dict)      # ficha_dict = asdict(FichaExpediente)
    con.commit()
    export_excel("data/proyectos.db", "proyectos.xlsx")

CLI:
    python store.py init      data/proyectos.db
    python store.py export    data/proyectos.db  proyectos.xlsx
    python store.py cargar    data/proyectos.db  ficha.json
"""
from __future__ import annotations

import argparse
import json
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

SCHEMA = Path(__file__).with_name("schema.sql")


def _ahora() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


# ────────────────────────────────────────────────────────────────────────────
# Conexión / init
# ────────────────────────────────────────────────────────────────────────────
def conectar(db_path: str | Path) -> sqlite3.Connection:
    db_path = Path(db_path)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(str(db_path))
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA foreign_keys = ON")
    con.executescript(SCHEMA.read_text(encoding="utf-8"))
    return con


# ────────────────────────────────────────────────────────────────────────────
# Upsert de una ficha
# ────────────────────────────────────────────────────────────────────────────
def _ultimo_movimiento(tramite: list[dict]) -> tuple[str | None, str | None]:
    """Devuelve (texto, fecha) del último movimiento con texto."""
    for mov in reversed(tramite or []):
        txt = (mov.get("movimiento") or "").strip()
        if txt:
            return txt, mov.get("fecha")
    return (None, None)


def upsert_proyecto(con: sqlite3.Connection, ficha: dict[str, Any]) -> str:
    """Inserta o actualiza un proyecto a partir de un dict FichaExpediente.

    Devuelve el denominador. NO hace commit (lo decide el llamador).
    """
    denom = ficha.get("expediente")
    if not denom:
        raise ValueError("La ficha no trae 'expediente' (denominador).")

    ahora = _ahora()
    tramite = ficha.get("tramite") or []
    ult_txt, ult_fecha = _ultimo_movimiento(tramite)

    fila = con.execute(
        "SELECT creado_en FROM proyectos WHERE denominador = ?", (denom,)
    ).fetchone()
    creado_en = fila["creado_en"] if fila and fila["creado_en"] else ahora

    con.execute(
        """
        INSERT INTO proyectos (denominador, camara, sumario, fecha_ingreso, estado,
            ultimo_movimiento, ultimo_movimiento_fecha, pdf_url, url, fuente_ok,
            capturado_en, creado_en, actualizado_en)
        VALUES (:denominador, :camara, :sumario, :fecha_ingreso, :estado,
            :ult, :ultf, :pdf_url, :url, :fuente_ok, :capturado_en, :creado_en, :actualizado_en)
        ON CONFLICT(denominador) DO UPDATE SET
            camara=excluded.camara, sumario=excluded.sumario,
            fecha_ingreso=excluded.fecha_ingreso, estado=excluded.estado,
            ultimo_movimiento=excluded.ultimo_movimiento,
            ultimo_movimiento_fecha=excluded.ultimo_movimiento_fecha,
            pdf_url=excluded.pdf_url, url=excluded.url, fuente_ok=excluded.fuente_ok,
            capturado_en=excluded.capturado_en, actualizado_en=excluded.actualizado_en
        """,
        {
            "denominador": denom,
            "camara": ficha.get("camara"),
            "sumario": ficha.get("sumario"),
            "fecha_ingreso": ficha.get("fecha_ingreso"),
            "estado": ficha.get("estado"),
            "ult": ult_txt,
            "ultf": ult_fecha,
            "pdf_url": ficha.get("pdf_url"),
            "url": ficha.get("url"),
            "fuente_ok": 1 if ficha.get("fuente_ok", True) else 0,
            "capturado_en": ficha.get("capturado_en"),
            "creado_en": creado_en,
            "actualizado_en": ahora,
        },
    )

    # Hijas que reflejan el estado oficial actual: se reemplazan completas.
    con.execute("DELETE FROM proyecto_autores WHERE denominador = ?", (denom,))
    for i, a in enumerate(ficha.get("firmantes") or []):
        con.execute(
            "INSERT INTO proyecto_autores (denominador, orden, nombre, distrito, bloque)"
            " VALUES (?,?,?,?,?)",
            (denom, i, a.get("nombre"), a.get("distrito"), a.get("bloque")),
        )

    con.execute("DELETE FROM proyecto_giros WHERE denominador = ?", (denom,))
    for g in ficha.get("giros") or []:
        con.execute(
            "INSERT INTO proyecto_giros (denominador, orden, comision,"
            " competencia_primaria, fecha_ingreso, fecha_egreso) VALUES (?,?,?,?,?,?)",
            (denom, g.get("orden"), g.get("comision"),
             1 if g.get("competencia_primaria") else 0,
             g.get("fecha_ingreso"), g.get("fecha_egreso")),
        )

    con.execute("DELETE FROM proyecto_tramite WHERE denominador = ?", (denom,))
    for i, m in enumerate(tramite):
        con.execute(
            "INSERT INTO proyecto_tramite (denominador, idx, camara, movimiento, fecha, resultado)"
            " VALUES (?,?,?,?,?,?)",
            (denom, i, m.get("camara"), m.get("movimiento"), m.get("fecha"), m.get("resultado")),
        )
    # taxonomias: intactas (las maneja el agente).
    return denom


def cargar_json(con: sqlite3.Connection, ruta: str | Path) -> str:
    data = json.loads(Path(ruta).read_text(encoding="utf-8"))
    if isinstance(data, list):
        ult = ""
        for ficha in data:
            ult = upsert_proyecto(con, ficha)
        return ult
    return upsert_proyecto(con, data)


# ────────────────────────────────────────────────────────────────────────────
# Export en formatos UNIVERSALES (sin separadores dentro de celdas).
#
# Regla de diseño: nada de "valor | valor | valor" en una celda. Lo multivaluado
# (autores, giros, trámite, taxonomías) va NORMALIZADO: una fila por valor, en su
# propia hoja/CSV, unido por `denominador`. Así cualquier herramienta (Excel,
# pandas, Power BI, R) lo puede cruzar sin parsear texto.
# ────────────────────────────────────────────────────────────────────────────

# (vista_nombre, SELECT). El orden de columnas lo define el propio SELECT.
_VISTAS: list[tuple[str, str]] = [
    ("Proyectos",
     "SELECT denominador, camara, fecha_ingreso, estado, ultimo_movimiento,"
     " ultimo_movimiento_fecha AS fecha_ult_mov, sumario, pdf_url, url AS ficha_url,"
     " fuente_ok, capturado_en, creado_en, actualizado_en"
     " FROM proyectos ORDER BY fecha_ingreso DESC, denominador"),
    ("Autores",
     "SELECT denominador, orden, nombre, distrito, bloque"
     " FROM proyecto_autores ORDER BY denominador, orden"),
    ("Giros",
     "SELECT denominador, orden, comision, competencia_primaria, fecha_ingreso, fecha_egreso"
     " FROM proyecto_giros ORDER BY denominador, orden"),
    ("Tramite",
     "SELECT denominador, idx, camara, movimiento, fecha, resultado"
     " FROM proyecto_tramite ORDER BY denominador, idx"),
    ("Taxonomias",
     "SELECT denominador, taxonomia_id, taxonomia, fuente, confianza, asignada_en"
     " FROM proyecto_taxonomias ORDER BY denominador, taxonomia_id"),
]


def _consulta(con: sqlite3.Connection, sql: str) -> tuple[list[str], list[list]]:
    cur = con.execute(sql)
    cols = [c[0] for c in cur.description]
    filas = [list(r) for r in cur.fetchall()]
    return cols, filas


def export_excel(db_path: str | Path, xlsx_path: str | Path) -> Path:
    """Workbook con una hoja por tabla (relacional, sin separadores)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    con = sqlite3.connect(str(db_path))
    wb = Workbook()
    wb.remove(wb.active)
    for nombre, sql in _VISTAS:
        ws = wb.create_sheet(nombre)
        cols, filas = _consulta(con, sql)
        ws.append(cols)
        for c in ws[1]:
            c.font = Font(bold=True)
        for fila in filas:
            ws.append(fila)
        for i, col in enumerate(cols, start=1):
            ancho = max(len(col) + 2, 12)
            if col in ("sumario", "movimiento"):
                ancho = 60
            elif col in ("pdf_url", "ficha_url"):
                ancho = 34
            elif col in ("nombre", "comision", "taxonomia"):
                ancho = 30
            ws.column_dimensions[get_column_letter(i)].width = ancho
        ws.freeze_panes = "A2"
        if filas:
            ws.auto_filter.ref = ws.dimensions
    con.close()
    out = Path(xlsx_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def export_csv(db_path: str | Path, out_dir: str | Path) -> list[Path]:
    """Un CSV por tabla (UTF-8 con BOM, comillas estándar). El formato más portable."""
    import csv

    con = sqlite3.connect(str(db_path))
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    escritos: list[Path] = []
    for nombre, sql in _VISTAS:
        cols, filas = _consulta(con, sql)
        ruta = out_dir / f"{nombre.lower()}.csv"
        # utf-8-sig = UTF-8 con BOM: Excel en español lo abre con acentos correctos.
        with ruta.open("w", encoding="utf-8-sig", newline="") as fh:
            w = csv.writer(fh)  # coma estándar, comillas automáticas
            w.writerow(cols)
            w.writerows(filas)
        escritos.append(ruta)
    con.close()
    return escritos


# ────────────────────────────────────────────────────────────────────────────
# CLI
# ────────────────────────────────────────────────────────────────────────────
def _main() -> None:
    p = argparse.ArgumentParser(description="Base de Proyectos de Ley (SQLite).")
    sub = p.add_subparsers(dest="cmd", required=True)
    pi = sub.add_parser("init", help="crea/inicializa la base")
    pi.add_argument("db")
    pc = sub.add_parser("cargar", help="upsert de una ficha JSON (objeto o lista)")
    pc.add_argument("db"); pc.add_argument("json")
    pe = sub.add_parser("export", help="exporta a Excel (una hoja por tabla)")
    pe.add_argument("db"); pe.add_argument("xlsx")
    pcsv = sub.add_parser("csv", help="exporta un CSV por tabla a una carpeta")
    pcsv.add_argument("db"); pcsv.add_argument("dir")

    args = p.parse_args()
    if args.cmd == "init":
        conectar(args.db).close()
        print(f"Base lista: {args.db}")
    elif args.cmd == "cargar":
        con = conectar(args.db)
        denom = cargar_json(con, args.json)
        con.commit(); con.close()
        print(f"Cargado: {denom}")
    elif args.cmd == "export":
        out = export_excel(args.db, args.xlsx)
        print(f"Excel: {out}")
    elif args.cmd == "csv":
        outs = export_csv(args.db, args.dir)
        print("CSV:", ", ".join(str(o) for o in outs))


if __name__ == "__main__":
    sys.exit(_main())
