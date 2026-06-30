"""Test de la base de Proyectos (SQLite). Sin red.

Corre: `python tests/test_store.py` (desde datos/proyectos/).
Valida: init, upsert, idempotencia (no duplica + actualiza estado),
preservación de taxonomías entre scrapes, y export a Excel.
"""
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))
import store  # noqa: E402


def _ok(cond, msg):
    print(("PASS" if cond else "FAIL"), "-", msg)
    assert cond, msg


# ficha tal como la entrega datos/seguimiento (asdict de FichaExpediente)
FICHA_V1 = {
    "expediente": "2832-D-2026", "camara": "diputados",
    "url": "https://www.hcdn.gov.ar/diputados/sajmechet/proyecto.html?exp=2832-D-2026",
    "sumario": "ACTOS DISCRIMINATORIOS - LEY 23592.", "fecha_ingreso": "2026-06-16",
    "firmantes": [
        {"nombre": "AJMECHET, SABRINA", "distrito": "CABA", "bloque": "LA LIBERTAD AVANZA"},
        {"nombre": "BANFI, KARINA", "distrito": "BUENOS AIRES", "bloque": "ADELANTE BUENOS AIRES"},
    ],
    "giros": [
        {"comision": "DERECHOS HUMANOS Y GARANTIAS", "orden": None,
         "competencia_primaria": True, "fecha_ingreso": None, "fecha_egreso": None},
        {"comision": "LEGISLACION PENAL", "orden": None,
         "competencia_primaria": False, "fecha_ingreso": None, "fecha_egreso": None},
    ],
    "tramite": [{"camara": "Diputados", "movimiento": "SOLICITUD DE COFIRMA", "fecha": None, "resultado": ""}],
    "pdf_url": "https://www.hcdn.gob.ar/proyectos/detalle_tp_adjunto/index.html?id=292542",
    "estado": "en_comision", "fuente_ok": True, "capturado_en": "2026-06-29T10:00:00+00:00",
}


def run():
    tmp = Path(tempfile.mkdtemp())
    db = tmp / "proyectos.db"

    con = store.conectar(db)
    store.upsert_proyecto(con, FICHA_V1)
    con.commit()

    n = con.execute("SELECT COUNT(*) FROM proyectos").fetchone()[0]
    _ok(n == 1, f"1 proyecto cargado ({n})")
    _ok(con.execute("SELECT COUNT(*) FROM proyecto_autores").fetchone()[0] == 2, "2 autores")
    _ok(con.execute("SELECT COUNT(*) FROM proyecto_giros").fetchone()[0] == 2, "2 giros")
    row = con.execute("SELECT estado, ultimo_movimiento, creado_en FROM proyectos").fetchone()
    _ok(row["estado"] == "en_comision", "estado v1")
    _ok(row["ultimo_movimiento"] == "SOLICITUD DE COFIRMA", "último movimiento derivado")
    creado_v1 = row["creado_en"]

    # el agente asigna una taxonomía (otra vía, no el scraper)
    con.execute(
        "INSERT INTO proyecto_taxonomias (denominador, taxonomia_id, taxonomia, fuente, confianza, asignada_en)"
        " VALUES (?,?,?,?,?,?)",
        ("2832-D-2026", "DDHH", "Derechos Humanos", "agente", 0.91, "2026-06-29T11:00:00+00:00"),
    )
    con.commit()

    # segundo scrape: el proyecto avanzó (media sanción) y sumó un firmante
    ficha_v2 = dict(FICHA_V1)
    ficha_v2["estado"] = "media_sancion"
    ficha_v2["tramite"] = FICHA_V1["tramite"] + [
        {"camara": "Diputados", "movimiento": "MEDIA SANCION", "fecha": "2026-07-01", "resultado": "APROBADO"}
    ]
    ficha_v2["firmantes"] = FICHA_V1["firmantes"] + [
        {"nombre": "RITONDO, CRISTIAN", "distrito": "BUENOS AIRES", "bloque": "PRO"}
    ]
    ficha_v2["capturado_en"] = "2026-07-01T10:00:00+00:00"
    store.upsert_proyecto(con, ficha_v2)
    con.commit()

    _ok(con.execute("SELECT COUNT(*) FROM proyectos").fetchone()[0] == 1,
        "sigue habiendo 1 proyecto (no duplicó)")
    row = con.execute("SELECT estado, ultimo_movimiento, creado_en FROM proyectos").fetchone()
    _ok(row["estado"] == "media_sancion", "estado actualizado a media_sancion")
    _ok(row["ultimo_movimiento"] == "MEDIA SANCION", "último movimiento actualizado")
    _ok(row["creado_en"] == creado_v1, "creado_en se preservó")
    _ok(con.execute("SELECT COUNT(*) FROM proyecto_autores").fetchone()[0] == 3,
        "autores refrescados (3)")
    _ok(con.execute("SELECT COUNT(*) FROM proyecto_taxonomias").fetchone()[0] == 1,
        "taxonomía del agente SOBREVIVIÓ al re-scrape")
    con.close()

    # export a Excel (una hoja por tabla, sin separadores)
    xlsx = tmp / "proyectos.xlsx"
    out = store.export_excel(db, xlsx)
    _ok(out.exists() and out.stat().st_size > 0, "Excel generado")
    from openpyxl import load_workbook
    wb = load_workbook(out)
    _ok(wb.sheetnames == ["Proyectos", "Autores", "Giros", "Tramite", "Taxonomias"],
        f"Excel: hojas relacionales ({wb.sheetnames})")
    wp = wb["Proyectos"]
    _ok(wp.max_row == 2, f"Excel Proyectos: encabezado + 1 fila ({wp.max_row})")
    fila = {h.value: c.value for h, c in zip(wp[1], wp[2])}
    _ok(fila["denominador"] == "2832-D-2026", "Excel: denominador")
    _ok(fila["estado"] == "media_sancion", "Excel: estado")
    _ok("|" not in str(wp[2][0].value), "Excel: sin separador pipe en celdas")
    wa = wb["Autores"]
    _ok(wa.max_row == 4, f"Excel Autores: 1 fila por autor ({wa.max_row-1} autores)")
    wt = wb["Taxonomias"]
    _ok(wt.max_row == 2 and wt["B2"].value == "DDHH", "Excel: taxonomía normalizada")

    # export a CSV universal (utf-8-sig)
    csv_dir = tmp / "csv"
    rutas = store.export_csv(db, csv_dir)
    nombres = sorted(r.name for r in rutas)
    _ok(nombres == ["autores.csv", "giros.csv", "proyectos.csv", "taxonomias.csv", "tramite.csv"],
        f"CSV: 5 archivos ({nombres})")
    import csv as _csv
    with (csv_dir / "autores.csv").open(encoding="utf-8-sig", newline="") as fh:
        rows = list(_csv.reader(fh))
    _ok(rows[0] == ["denominador", "orden", "nombre", "distrito", "bloque"], "CSV: encabezados autores")
    _ok(len(rows) == 4, f"CSV autores: header + 3 filas ({len(rows)})")
    print("\nTODOS LOS TESTS PASARON")


if __name__ == "__main__":
    run()
